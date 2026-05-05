import os
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook

DATA_DIR = ".github/datafiles"
TEMPLATE_FILE = "template.html"
OUTPUT_FILE = "index.html"

def load_mapping():
    # Based on the analysis, we'll use a fixed mapping for the sheets we found
    return {
        "Target Date": "Target Date",
        "Description": "Description",
        "Type": "Type",
        "Producer": "Producer",
        "Audience": "Audience",
        "BBS Job Number": "BBS Job Number",
        "Vanity Link": "Vanity Link",
        "Premium": "Premium",
        "BBS Final Art": "BBS Final Art",
        "Segment Code": "Segment Code"
    }

def extract_hyperlinks(file_path, sheet_name, mapping):
    links = {}
    if not file_path.endswith(('.xlsx', '.xls')):
        return links
    
    try:
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return links
        ws = wb[sheet_name]
        
        headers = [cell.value for cell in ws[1]]
        target_cols = {}
        for field in ["Vanity Link", "Premium", "BBS Final Art"]:
            col_name = mapping.get(field, field)
            if col_name in headers:
                target_cols[field] = headers.index(col_name) + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_links = {}
            for field, col_idx in target_cols.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.hyperlink:
                    row_links[field] = cell.hyperlink.target
            if row_links:
                links[row_idx] = row_links
    except Exception as e:
        print(f"Error extracting links from {file_path} [{sheet_name}]: {e}")
    return links

def process_file(file_path, mapping):
    all_events = []
    all_standing = []
    all_tapings = []

    try:
        if file_path.endswith(('.xlsx', '.xls')):
            xl = pd.ExcelFile(file_path)
            for sheet in xl.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet)
                
                if "Standard" in sheet or "Standing" in sheet:
                    # Normalize standing data
                    for _, row in df.iterrows():
                        if pd.isna(row.get('Name of Communication')): continue
                        all_standing.append({
                            "name": str(row.get('Name of Communication', '')).strip(),
                            "abbreviation": str(row.get('Nickname / Abbreviation', '')).strip() if not pd.isna(row.get('Nickname / Abbreviation')) else "",
                            "type": str(row.get('Type', '')).strip(),
                            "frequency": str(row.get('How Often', '')).strip(),
                            "audience": str(row.get('Audience', '')).strip(),
                            "subject": str(row.get('Subject', '')).strip() if not pd.isna(row.get('Subject')) else "",
                            "notes": str(row.get('Notes', '')).strip() if not pd.isna(row.get('Notes')) else ""
                        })
                elif "Taping" in sheet:
                    # Normalize tapings data
                    # The tapings sheet seems to have dates as headers or single column
                    for col in df.columns:
                        if 'April' in col or 'August' in col or 'Sept' in col:
                            all_tapings.append({"label": col, "start": col, "end": col})
                        for val in df[col].dropna():
                            val_str = str(val)
                            if any(m in val_str for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                                all_tapings.append({"label": val_str, "start": val_str, "end": val_str})
                else:
                    links = extract_hyperlinks(file_path, sheet, mapping)
                    events = df.to_dict('records')
                    for i, ev in enumerate(events):
                        if (i + 2) in links:
                            ev['_links'] = links[i + 2]
                    all_events.extend(events)
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
            all_events.extend(df.to_dict('records'))
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
    
    return all_events, all_standing, all_tapings

def normalize_events(raw_events, mapping):
    normalized = []
    date_col = mapping.get("Target Date", "Target Date")
    desc_col = mapping.get("Description", "Description")
    type_col = mapping.get("Type", "Type")
    prod_col = mapping.get("Producer", "Producer")
    aud_col  = mapping.get("Audience", "Audience")
    job_col  = mapping.get("BBS Job Number", "BBS Job Number")
    van_col  = mapping.get("Vanity Link", "Vanity Link")
    prem_col = mapping.get("Premium", "Premium")
    art_col  = mapping.get("BBS Final Art", "BBS Final Art")
    seg_col  = mapping.get("Segment Code", "Segment Code")

    for ev in raw_events:
        d_val = ev.get(date_col)
        if pd.isna(d_val): continue
        
        try:
            if isinstance(d_val, datetime):
                dt = d_val
            else:
                dt = pd.to_datetime(str(d_val))
            date_str = dt.strftime("%Y-%m-%d")
        except:
            continue

        links = ev.get('_links', {})
        
        normalized.append({
            "date": date_str,
            "title": str(ev.get(desc_col, "")).strip(),
            "type": str(ev.get(type_col, "")).strip(),
            "producer": str(ev.get(prod_col, "")).strip(),
            "audience": str(ev.get(aud_col, "")).strip(),
            "job": str(ev.get(job_col, "")).strip() if not pd.isna(ev.get(job_col)) else "",
            "vanity": str(ev.get(van_col, "")).strip(),
            "vanity_url": links.get("Vanity Link"),
            "premium": str(ev.get(prem_col, "")).strip(),
            "premium_url": links.get("Premium"),
            "premium_on": str(ev.get(prem_col, "")).lower() not in ["", "nan", "none", "—"],
            "art": str(ev.get(art_col, "")).strip(),
            "art_url": links.get("BBS Final Art"),
            "segment": str(ev.get(seg_col, "")).strip()
        })
    
    return sorted(normalized, key=lambda x: x['date'])

def main():
    mapping = load_mapping()
    files = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) if f.endswith(('.xlsx', '.xls', '.csv'))]
    
    all_raw_events = []
    all_standing = []
    all_tapings = []
    
    for f in files:
        print(f"Processing {f}...")
        e, s, t = process_file(f, mapping)
        all_raw_events.extend(e)
        all_standing.extend(s)
        all_tapings.extend(t)
        
    events = normalize_events(all_raw_events, mapping)
    
    payload = {
        "events": events,
        "standing": all_standing,
        "tapings": all_tapings,
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M")
    }

    if not os.path.exists(TEMPLATE_FILE):
        print(f"Template {TEMPLATE_FILE} not found!")
        return

    with open(TEMPLATE_FILE, 'r') as f:
        html = f.read()
    
    # Inject payload
    payload_json = json.dumps(payload)
    html = html.replace("const EV = [];", f"const EV = {json.dumps(events)};")
    html = html.replace("const STANDING = [];", f"const STANDING = {json.dumps(all_standing)};")
    html = html.replace("const TAPINGS = [];", f"const TAPINGS = {json.dumps(all_tapings)};")
    # Also handle the generic placeholder if it exists
    html = html.replace("__PAYLOAD_PLACEHOLDER__", payload_json)
    
    with open(OUTPUT_FILE, 'w') as f:
        f.write(html)
    
    print(f"Successfully generated {OUTPUT_FILE}")
    print(f"Events: {len(events)}, Standing: {len(all_standing)}, Tapings: {len(all_tapings)}")

if __name__ == "__main__":
    main()
